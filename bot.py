import re
import io
import sqlite3
import logging
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from datetime import datetime, date
import os

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

TOKEN = os.environ.get("TOKEN", "SEU_TOKEN_AQUI")

logging.basicConfig(
    format="%(asctime)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

CATEGORIAS_PADRAO = [
    "Alimentação", "Transporte", "Saúde", "Moradia",
    "Lazer", "Educação", "Vestuário", "Outros"
]

KEYWORDS = {
    "Alimentação":  ["mercado", "supermercado", "restaurante", "lanche", "comida",
                     "padaria", "pizza", "hamburguer", "ifood", "almoço", "jantar", "café"],
    "Transporte":   ["uber", "99", "gasolina", "combustivel", "ônibus", "metro",
                     "estacionamento", "pedagio", "taxi", "passagem"],
    "Saúde":        ["farmacia", "remedio", "medico", "consulta", "exame",
                     "hospital", "dentista", "academia", "plano de saude"],
    "Moradia":      ["aluguel", "condominio", "agua", "luz", "energia", "internet",
                     "telefone", "conta", "boleto", "gas"],
    "Lazer":        ["cinema", "teatro", "show", "viagem", "hotel", "bar",
                     "festa", "streaming", "netflix", "spotify", "jogo"],
    "Educação":     ["curso", "livro", "escola", "faculdade", "mensalidade",
                     "material", "aula", "treinamento"],
    "Vestuário":    ["roupa", "sapato", "tenis", "calcado", "loja", "shopping"],
}

CORES_CATEGORIA = {
    "Alimentação": "#FF6B6B",
    "Transporte":  "#4ECDC4",
    "Saúde":       "#45B7D1",
    "Moradia":     "#96CEB4",
    "Lazer":       "#FFEAA7",
    "Educação":    "#DDA0DD",
    "Vestuário":   "#98D8C8",
    "Outros":      "#B0B0B0",
}

def init_db():
    con = sqlite3.connect("financas.db")
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS gastos (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            chat_id    INTEGER NOT NULL,
            valor      REAL    NOT NULL,
            descricao  TEXT    NOT NULL,
            categoria  TEXT    NOT NULL,
            data       TEXT    NOT NULL,
            mes        TEXT    NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS categorias_custom (
            chat_id   INTEGER NOT NULL,
            categoria TEXT    NOT NULL,
            PRIMARY KEY (chat_id, categoria)
        )
    """)
    con.commit()
    con.close()

def db():
    return sqlite3.connect("financas.db")

def mes_atual():
    return date.today().strftime("%Y-%m")

def hoje():
    return date.today().strftime("%Y-%m-%d")

def get_categorias(chat_id):
    con = db()
    extras = [r[0] for r in con.execute(
        "SELECT categoria FROM categorias_custom WHERE chat_id=?", (chat_id,)
    ).fetchall()]
    con.close()
    return CATEGORIAS_PADRAO + extras

def detectar_categoria(descricao):
    desc = descricao.lower()
    for cat, palavras in KEYWORDS.items():
        if any(p in desc for p in palavras):
            return cat
    return "Outros"

def extrair_gasto(texto):
    texto = texto.strip()
    padroes = [
        r'(?:gastei|paguei|comprei|gasto|despesa|saiu|saída)\s+r?\$?\s*([\d]+[.,]?[\d]*)\s+(?:de\s+|no\s+|na\s+|em\s+|com\s+)?(.*)',
        r'r?\$?\s*([\d]+[.,]?[\d]*)\s+(?:de\s+|no\s+|na\s+|em\s+|com\s+)?(.*)',
        r'(.*?)\s+r?\$?\s*([\d]+[.,]?[\d]*)',
    ]
    for i, padrao in enumerate(padroes):
        m = re.match(padrao, texto, re.IGNORECASE)
        if m:
            if i < 2:
                valor_str = m.group(1).replace(",", ".")
                descricao = m.group(2).strip() if m.group(2) else "Gasto"
            else:
                descricao = m.group(1).strip() if m.group(1) else "Gasto"
                valor_str = m.group(2).replace(",", ".")
            try:
                valor = float(valor_str)
                if valor > 0:
                    return valor, descricao or "Gasto"
            except:
                pass
    return None, None

def teclado_categorias(chat_id, prefixo="cat"):
    cats = get_categorias(chat_id)
    botoes = []
    linha = []
    for i, cat in enumerate(cats):
        linha.append(InlineKeyboardButton(cat, callback_data=f"{prefixo}:{cat}"))
        if len(linha) == 2:
            botoes.append(linha)
            linha = []
    if linha:
        botoes.append(linha)
    return InlineKeyboardMarkup(botoes)

PENDENTES = {}

def menu_principal():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Registrar Gasto",  callback_data="menu_registrar"),
         InlineKeyboardButton("📊 Relatório do Mês", callback_data="menu_relatorio")],
        [InlineKeyboardButton("📈 Gráfico de Pizza", callback_data="menu_pizza"),
         InlineKeyboardButton("📉 Gráfico Mensal",   callback_data="menu_mensal")],
        [InlineKeyboardButton("📋 Últimos Gastos",   callback_data="menu_ultimos"),
         InlineKeyboardButton("🏷 Categorias",       callback_data="menu_categorias")],
        [InlineKeyboardButton("📥 Exportar Excel",   callback_data="menu_excel"),
         InlineKeyboardButton("🗑 Deletar Gasto",    callback_data="menu_deletar")],
        [InlineKeyboardButton("ℹ️ Ajuda",            callback_data="menu_ajuda")],
    ])

async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    nome = update.effective_user.first_name
    await update.message.reply_text(
        f"Olá, {nome}! 💰 Sou seu assistente financeiro pessoal.\n\n"
        "Pode me dizer seus gastos naturalmente:\n"
        "• _gastei 50 no mercado_\n"
        "• _paguei 120 de aluguel_\n"
        "• _uber 22,50_\n\n"
        "Ou use os comandos abaixo 👇",
        parse_mode="Markdown",
        reply_markup=menu_principal()
    )

async def cmd_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("O que deseja fazer?", reply_markup=menu_principal())

async def mensagem_livre(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text
    chat_id = update.effective_chat.id
    valor, descricao = extrair_gasto(texto)
    if valor is None:
        await update.message.reply_text(
            "Não entendi o valor. Tente assim:\n"
            "• _gastei 45 no mercado_\n"
            "• _paguei 30 de uber_\n"
            "• _120 aluguel_\n\n"
            "Ou use /menu para navegar.",
            parse_mode="Markdown"
        )
        return
    categoria_auto = detectar_categoria(descricao)
    PENDENTES[chat_id] = {"valor": valor, "descricao": descricao}
    await update.message.reply_text(
        f"Entendi! 💸\n\n"
        f"*Valor:* R$ {valor:.2f}\n"
        f"*Descrição:* {descricao}\n"
        f"*Categoria sugerida:* {categoria_auto}\n\n"
        "Confirma a categoria ou escolha outra:",
        parse_mode="Markdown",
        reply_markup=teclado_categorias(chat_id, prefixo="confirmar")
    )

async def callback_confirmar_categoria(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    chat_id = query.message.chat_id
    await query.answer()
    if chat_id not in PENDENTES:
        await query.message.reply_text("Sessão expirada. Digite o gasto novamente.")
        return
    categoria = query.data.split(":", 1)[1]
    dados = PENDENTES.pop(chat_id)
    valor = dados["valor"]
    descricao = dados["descricao"]
    con = db()
    con.execute(
        "INSERT INTO gastos (chat_id, valor, descricao, categoria, data, mes) VALUES (?,?,?,?,?,?)",
        (chat_id, valor, descricao, categoria, hoje(), mes_atual())
    )
    con.commit()
    con.close()
    await query.message.reply_text(
        f"✅ Gasto registrado!\n\n"
        f"💸 *R$ {valor:.2f}* — {descricao}\n"
        f"🏷 Categoria: *{categoria}*\n"
        f"📅 Data: {hoje()}",
        parse_mode="Markdown",
        reply_markup=menu_principal()
    )

async def relatorio(update: Update, ctx: ContextTypes.DEFAULT_TYPE, chat_id=None, msg=None):
    chat_id = chat_id or update.effective_chat.id
    msg = msg or update.message
    mes = mes_atual()
    con = db()
    gastos = con.execute(
        "SELECT categoria, SUM(valor) FROM gastos WHERE chat_id=? AND mes=? GROUP BY categoria ORDER BY SUM(valor) DESC",
        (chat_id, mes)
    ).fetchall()
    total = con.execute(
        "SELECT SUM(valor) FROM gastos WHERE chat_id=? AND mes=?",
        (chat_id, mes)
    ).fetchone()[0] or 0
    con.close()
    if not gastos:
        await msg.reply_text("📭 Nenhum gasto registrado este mês ainda.")
        return
    ano, m = mes.split("-")
    nome_mes = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"][int(m)-1]
    linhas = [f"📊 *Relatório — {nome_mes}/{ano}*\n"]
    for cat, val in gastos:
        pct = (val / total * 100) if total else 0
        linhas.append(f"  {cat}: *R$ {val:.2f}* ({pct:.1f}%)")
    linhas.append(f"\n💰 *Total: R$ {total:.2f}*")
    await msg.reply_text("\n".join(linhas), parse_mode="Markdown", reply_markup=menu_principal())

async def grafico_pizza(update: Update, ctx: ContextTypes.DEFAULT_TYPE, chat_id=None, msg=None):
    chat_id = chat_id or update.effective_chat.id
    msg = msg or update.message
    mes = mes_atual()
    con = db()
    gastos = con.execute(
        "SELECT categoria, SUM(valor) FROM gastos WHERE chat_id=? AND mes=? GROUP BY categoria ORDER BY SUM(valor) DESC",
        (chat_id, mes)
    ).fetchall()
    total = con.execute(
        "SELECT SUM(valor) FROM gastos WHERE chat_id=? AND mes=?",
        (chat_id, mes)
    ).fetchone()[0] or 0
    con.close()
    if not gastos:
        await msg.reply_text("📭 Nenhum gasto este mês para gerar gráfico.")
        return
    cats = [g[0] for g in gastos]
    vals = [g[1] for g in gastos]
    cores = [CORES_CATEGORIA.get(c, "#CCCCCC") for c in cats]
    fig, ax = plt.subplots(figsize=(8, 6), facecolor="#1a1a2e")
    ax.set_facecolor("#1a1a2e")
    wedges, texts, autotexts = ax.pie(
        vals, labels=None, autopct="%1.1f%%",
        colors=cores, startangle=140, pctdistance=0.75,
        wedgeprops=dict(width=0.6, edgecolor="#1a1a2e", linewidth=2)
    )
    for at in autotexts:
        at.set_color("white")
        at.set_fontsize(9)
        at.set_fontweight("bold")
    legenda = [mpatches.Patch(color=c, label=f"{cat}  R$ {val:.2f}")
               for cat, val, c in zip(cats, vals, cores)]
    ax.legend(handles=legenda, loc="lower center", bbox_to_anchor=(0.5, -0.18),
              ncol=2, framealpha=0, labelcolor="white", fontsize=9)
    mes_nome = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"][int(mes.split("-")[1])-1]
    ax.set_title(f"Gastos por Categoria — {mes_nome}/{mes.split('-')[0]}\nTotal: R$ {total:.2f}",
                 color="white", fontsize=12, fontweight="bold", pad=20)
    buf = io.BytesIO()
    plt.savefig(buf, format="png", bbox_inches="tight", facecolor="#1a1a2e")
    buf.seek(0)
    plt.close()
    await msg.reply_photo(photo=buf, reply_markup=menu_principal())

async def grafico_mensal(update: Update, ctx: ContextTypes.DEFAULT_TYPE, chat_id=None, msg=None):
    chat_id = chat_id or update.effective_chat.id
    msg = msg or update.message
    con = db()
    dados = con.execute(
        "SELECT mes, SUM(valor) FROM gastos WHERE chat_id=? GROUP BY mes ORDER BY mes",
        (chat_id,)
    ).fetchall()
    con.close()
    if not dados:
        await msg.reply_text("📭 Nenhum dado histórico para gerar gráfico.")
        return
    meses = [d[0] for d in dados]
    totais = [d[1] for d in dados]
    labels = [["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"][int(m.split("-")[1])-1]
              + f"\n{m.split('-')[0]}" for m in meses]
    fig, ax = plt.subplots(figsize=(max(6, len(meses)*1.2), 5), facecolor="#1a1a2e")
    ax.set_facecolor("#16213e")
    bars = ax.bar(labels, totais, color="#4ECDC4", edgecolor="#1a1a2e", linewidth=1.5, zorder=3)
    for bar, val in zip(bars, totais):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(totais)*0.01,
                f"R$ {val:.0f}", ha="center", va="bottom", color="white",
                fontsize=8, fontweight="bold")
    ax.set_title("Gastos por Mês", color="white", fontsize=13, fontweight="bold", pad=15)
    ax.set_ylabel("R$", color="white")
    ax.tick_params(colors="white")
    ax.spines[["top","right","left","bottom"]].set_visible(False)
    ax.yaxis.grid(True, color="#ffffff22", zorder=0)
    ax.set_axisbelow(True)
    buf = io.BytesIO()
    plt.savefig(buf, format="png", bbox_inches="tight", facecolor="#1a1a2e")
    buf.seek(0)
    plt.close()
    await msg.reply_photo(photo=buf, reply_markup=menu_principal())

async def ultimos_gastos(update: Update, ctx: ContextTypes.DEFAULT_TYPE, chat_id=None, msg=None):
    chat_id = chat_id or update.effective_chat.id
    msg = msg or update.message
    con = db()
    gastos = con.execute(
        "SELECT id, valor, descricao, categoria, data FROM gastos WHERE chat_id=? ORDER BY id DESC LIMIT 10",
        (chat_id,)
    ).fetchall()
    con.close()
    if not gastos:
        await msg.reply_text("📭 Nenhum gasto registrado ainda.")
        return
    linhas = ["📋 *Últimos 10 gastos:*\n"]
    for gid, val, desc, cat, data in gastos:
        linhas.append(f"*#{gid}* | R$ {val:.2f} | {desc} | _{cat}_ | {data}")
    linhas.append("\n🗑 Para deletar: /deletar <id>")
    await msg.reply_text("\n".join(linhas), parse_mode="Markdown", reply_markup=menu_principal())

async def exportar_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE, chat_id=None, msg=None):
    chat_id = chat_id or update.effective_chat.id
    msg = msg or update.message

    if not EXCEL_OK:
        await msg.reply_text("❌ Biblioteca Excel não disponível no servidor.")
        return

    con = db()
    todos = con.execute(
        "SELECT data, descricao, categoria, valor, mes FROM gastos WHERE chat_id=? ORDER BY data DESC",
        (chat_id,)
    ).fetchall()
    resumo_cat = con.execute(
        "SELECT categoria, SUM(valor) FROM gastos WHERE chat_id=? AND mes=? GROUP BY categoria ORDER BY SUM(valor) DESC",
        (chat_id, mes_atual())
    ).fetchall()
    resumo_mes = con.execute(
        "SELECT mes, SUM(valor) FROM gastos WHERE chat_id=? GROUP BY mes ORDER BY mes",
        (chat_id,)
    ).fetchall()
    con.close()

    if not todos:
        await msg.reply_text("📭 Nenhum gasto registrado para exportar.")
        return

    def borda():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def cel(ws, row, col, valor=None, bold=False, bg=None, fg="000000", fmt=None, align="center"):
        c = ws.cell(row=row, column=col)
        if valor is not None:
            c.value = valor
        c.font = Font(name="Arial", bold=bold, color=fg, size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = borda()
        if bg:
            c.fill = PatternFill("solid", start_color=bg)
        if fmt:
            c.number_format = fmt
        return c

    wb = openpyxl.Workbook()

    # ── Aba 1: Lançamentos ──
    ws1 = wb.active
    ws1.title = "Lançamentos"
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 32
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 12

    for i, cab in enumerate(["Data", "Descrição", "Categoria", "Valor (R$)", "Mês"], 1):
        cel(ws1, 1, i, cab, bold=True, bg="1F4E79", fg="FFFFFF")

    for r, (data, desc, cat, val, mes) in enumerate(todos, 2):
        bg = "F7F7F7" if r % 2 == 0 else "FFFFFF"
        cel(ws1, r, 1, data, bg=bg, align="center")
        cel(ws1, r, 2, desc, bg=bg, align="left")
        cel(ws1, r, 3, cat,  bg=bg, align="center")
        cel(ws1, r, 4, val,  bg=bg, fmt='"R$"#,##0.00')
        cel(ws1, r, 5, mes,  bg=bg, align="center")

    # ── Aba 2: Resumo por Categoria ──
    ws2 = wb.create_sheet("Resumo por Categoria")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 12

    total_mes = sum(v for _, v in resumo_cat)
    mes_nome = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"][int(mes_atual().split("-")[1])-1]
    ws2.merge_cells("A1:C1")
    t = ws2["A1"]
    t.value = f"Resumo do Mês — {mes_nome}/{mes_atual().split('-')[0]}"
    t.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    for i, cab in enumerate(["Categoria", "Total (R$)", "% do Mês"], 1):
        cel(ws2, 2, i, cab, bold=True, bg="2E75B6", fg="FFFFFF")

    for r, (cat, val) in enumerate(resumo_cat, 3):
        pct = (val / total_mes * 100) if total_mes else 0
        bg = "F7F7F7" if r % 2 == 0 else "FFFFFF"
        cel(ws2, r, 1, cat,  bg=bg, align="left")
        cel(ws2, r, 2, val,  bg=bg, fmt='"R$"#,##0.00')
        cel(ws2, r, 3, pct/100, bg=bg, fmt='0.0%')

    r_tot = len(resumo_cat) + 3
    cel(ws2, r_tot, 1, "TOTAL", bold=True, bg="70AD47", fg="FFFFFF")
    cel(ws2, r_tot, 2, total_mes, bold=True, bg="70AD47", fg="FFFFFF", fmt='"R$"#,##0.00')
    cel(ws2, r_tot, 3, 1.0, bold=True, bg="70AD47", fg="FFFFFF", fmt='0.0%')

    # ── Aba 3: Histórico Mensal ──
    ws3 = wb.create_sheet("Histórico Mensal")
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 18

    for i, cab in enumerate(["Mês", "Total (R$)"], 1):
        cel(ws3, 1, i, cab, bold=True, bg="1F4E79", fg="FFFFFF")

    for r, (mes, val) in enumerate(resumo_mes, 2):
        ano_m, m_m = mes.split("-")
        nome = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"][int(m_m)-1]
        bg = "F7F7F7" if r % 2 == 0 else "FFFFFF"
        cel(ws3, r, 1, f"{nome}/{ano_m}", bg=bg)
        cel(ws3, r, 2, val, bg=bg, fmt='"R$"#,##0.00')

    # Salvar em memória e enviar
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "financas.xlsx"

    await msg.reply_document(
        document=buf,
        filename=f"financas_{hoje()}.xlsx",
        caption="📥 Aqui está sua planilha com todos os gastos!",
        reply_markup=menu_principal()
    )

async def deletar_gasto(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not ctx.args or not ctx.args[0].isdigit():
        await update.message.reply_text("Use: /deletar <id>\nVeja os IDs em /ultimos")
        return
    gid = int(ctx.args[0])
    con = db()
    cur = con.execute("DELETE FROM gastos WHERE id=? AND chat_id=?", (gid, chat_id))
    con.commit()
    con.close()
    if cur.rowcount:
        await update.message.reply_text(f"🗑 Gasto #{gid} deletado.")
    else:
        await update.message.reply_text(f"❌ Gasto #{gid} não encontrado.")

async def ver_categorias(update: Update, ctx: ContextTypes.DEFAULT_TYPE, chat_id=None, msg=None):
    chat_id = chat_id or update.effective_chat.id
    msg = msg or update.message
    cats = get_categorias(chat_id)
    texto = "🏷 *Categorias disponíveis:*\n\n" + "\n".join(f"• {c}" for c in cats)
    texto += "\n\nPara adicionar: /addcategoria <nome>"
    await msg.reply_text(texto, parse_mode="Markdown", reply_markup=menu_principal())

async def add_categoria(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not ctx.args:
        await update.message.reply_text("Use: /addcategoria <nome>")
        return
    nova = " ".join(ctx.args).strip().title()
    if nova in get_categorias(chat_id):
        await update.message.reply_text(f"A categoria *{nova}* já existe.", parse_mode="Markdown")
        return
    con = db()
    con.execute("INSERT OR IGNORE INTO categorias_custom (chat_id, categoria) VALUES (?,?)", (chat_id, nova))
    con.commit()
    con.close()
    await update.message.reply_text(f"✅ Categoria *{nova}* adicionada!", parse_mode="Markdown")

async def callback_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    chat_id = query.message.chat_id
    await query.answer()
    msg = query.message
    acao = query.data
    if acao == "menu_registrar":
        await msg.reply_text(
            "💬 Digite seu gasto no chat:\n\n"
            "• _gastei 50 no mercado_\n"
            "• _paguei 35 de uber_\n"
            "• _120,50 conta de luz_",
            parse_mode="Markdown"
        )
    elif acao == "menu_relatorio":
        await relatorio(update, ctx, chat_id=chat_id, msg=msg)
    elif acao == "menu_pizza":
        await grafico_pizza(update, ctx, chat_id=chat_id, msg=msg)
    elif acao == "menu_mensal":
        await grafico_mensal(update, ctx, chat_id=chat_id, msg=msg)
    elif acao == "menu_ultimos":
        await ultimos_gastos(update, ctx, chat_id=chat_id, msg=msg)
    elif acao == "menu_categorias":
        await ver_categorias(update, ctx, chat_id=chat_id, msg=msg)
    elif acao == "menu_excel":
        await exportar_excel(update, ctx, chat_id=chat_id, msg=msg)
    elif acao == "menu_deletar":
        await msg.reply_text("Para deletar um gasto:\n/deletar <id>\n\nVeja os IDs com /ultimos")
    elif acao == "menu_ajuda":
        await msg.reply_text(
            "ℹ️ *Como usar o assistente:*\n\n"
            "📝 *Registrar gasto:* apenas digite no chat\n"
            "Ex: _gastei 50 no mercado_\n\n"
            "📊 *Relatório:* resumo por categoria do mês\n"
            "📈 *Pizza:* gráfico visual de categorias\n"
            "📉 *Mensal:* evolução mês a mês\n"
            "📋 *Últimos:* últimos 10 lançamentos\n"
            "📥 *Exportar Excel:* baixar planilha completa\n"
            "🏷 *Categorias:* ver e adicionar categorias\n\n"
            "*Comandos:*\n"
            "/menu — Menu principal\n"
            "/relatorio — Relatório do mês\n"
            "/ultimos — Últimos gastos\n"
            "/deletar <id> — Remover lançamento\n"
            "/addcategoria <nome> — Nova categoria\n"
            "/excel — Exportar planilha",
            parse_mode="Markdown",
            reply_markup=menu_principal()
        )
    elif acao.startswith("confirmar:"):
        await callback_confirmar_categoria(update, ctx)

def main():
    init_db()
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start",        start))
    app.add_handler(CommandHandler("menu",         cmd_menu))
    app.add_handler(CommandHandler("relatorio",    relatorio))
    app.add_handler(CommandHandler("ultimos",      ultimos_gastos))
    app.add_handler(CommandHandler("deletar",      deletar_gasto))
    app.add_handler(CommandHandler("addcategoria", add_categoria))
    app.add_handler(CommandHandler("categorias",   ver_categorias))
    app.add_handler(CommandHandler("excel",        exportar_excel))
    app.add_handler(CallbackQueryHandler(callback_confirmar_categoria, pattern="^confirmar:"))
    app.add_handler(CallbackQueryHandler(callback_menu))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, mensagem_livre))
    print("💰 Bot financeiro rodando!")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
